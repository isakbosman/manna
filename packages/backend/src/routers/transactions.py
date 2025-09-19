"""
Transaction management router for Manna Financial Platform.
Handles transaction listing, filtering, updates, bulk operations, and exports.
"""

from fastapi import APIRouter, Depends, HTTPException, Query, File, UploadFile, BackgroundTasks
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_, func, extract
from typing import List, Optional, Dict, Any
from datetime import date, datetime
from uuid import UUID
import csv
import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..database import get_db
from ..database.models import Transaction, Account, Category
from ..schemas.transaction import (
    Transaction as TransactionSchema,
    TransactionUpdate,
    TransactionList,
    TransactionFilter,
    TransactionExport,
    BulkTransactionUpdate,
    TransactionStats,
    CategorySummary
)
from ..schemas.common import PaginatedResponse
from ..dependencies.auth import get_current_verified_user
from ..database.models import User
# from ..utils.export import generate_csv, generate_excel  # TODO: Implement export utilities

router = APIRouter()


@router.get("/", response_model=PaginatedResponse[TransactionSchema])
async def list_transactions(
    # Pagination
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(50, ge=1, le=500, description="Items per page"),
    
    # Date filters
    start_date: Optional[date] = Query(None, description="Start date filter"),
    end_date: Optional[date] = Query(None, description="End date filter"),
    
    # Amount filters
    min_amount: Optional[float] = Query(None, description="Minimum amount"),
    max_amount: Optional[float] = Query(None, description="Maximum amount"),
    
    # Category filters
    category: Optional[str] = Query(None, description="Primary category filter"),
    detailed_category: Optional[str] = Query(None, description="Detailed category filter"),
    
    # Account filters
    account_id: Optional[UUID] = Query(None, description="Filter by account"),
    account_type: Optional[str] = Query(None, description="Filter by account type"),
    
    # Status filters
    pending: Optional[bool] = Query(None, description="Filter pending transactions"),
    reconciled: Optional[bool] = Query(None, description="Filter reconciled transactions"),
    
    # Search
    search: Optional[str] = Query(None, description="Search in transaction name/merchant"),
    
    # Sorting
    sort_by: str = Query("date", description="Sort field"),
    sort_order: str = Query("desc", description="Sort order (asc/desc)"),
    
    # Dependencies
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_verified_user)
):
    """
    List transactions with advanced filtering, search, and pagination.
    """
    # Base query - only user's transactions
    query = db.query(Transaction).join(Account).filter(Account.user_id == current_user.id)
    
    # Apply filters
    filters = []
    
    if start_date:
        filters.append(Transaction.date >= start_date)
    if end_date:
        filters.append(Transaction.date <= end_date)
    
    if min_amount is not None:
        filters.append(Transaction.amount_cents >= int(min_amount * 100))
    if max_amount is not None:
        filters.append(Transaction.amount_cents <= int(max_amount * 100))
    
    if category:
        filters.append(Transaction.primary_category == category)
    if detailed_category:
        filters.append(Transaction.detailed_category == detailed_category)
    
    if account_id:
        filters.append(Transaction.account_id == account_id)
    if account_type:
        filters.append(Account.type == account_type)
    
    if pending is not None:
        filters.append(Transaction.pending == pending)
    if reconciled is not None:
        filters.append(Transaction.is_reconciled == reconciled)
    
    if search:
        search_term = f"%{search}%"
        filters.append(
            or_(
                Transaction.name.ilike(search_term),
                Transaction.merchant_name.ilike(search_term),
                Transaction.original_description.ilike(search_term)
            )
        )
    
    if filters:
        query = query.filter(and_(*filters))
    
    # Get total count
    total_items = query.count()
    
    # Apply sorting
    sort_column = getattr(Transaction, sort_by, Transaction.date)
    if sort_order.lower() == "desc":
        query = query.order_by(sort_column.desc())
    else:
        query = query.order_by(sort_column.asc())
    
    # Apply pagination
    offset = (page - 1) * page_size
    transactions = query.offset(offset).limit(page_size).all()
    
    # Calculate pagination info
    total_pages = (total_items + page_size - 1) // page_size
    
    return PaginatedResponse(
        items=transactions,
        total=total_items,
        page=page,
        per_page=page_size,
        pages=total_pages
    )


@router.get("/stats", response_model=TransactionStats)
async def get_transaction_stats(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    account_id: Optional[UUID] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_verified_user)
):
    """
    Get transaction statistics and summaries.
    """
    query = db.query(Transaction).join(Account).filter(Account.user_id == current_user.id)
    
    if start_date:
        query = query.filter(Transaction.date >= start_date)
    if end_date:
        query = query.filter(Transaction.date <= end_date)
    if account_id:
        query = query.filter(Transaction.account_id == account_id)
    
    # Calculate statistics
    total_count = query.count()
    total_income = query.filter(Transaction.amount > 0).with_entities(
        func.sum(Transaction.amount)
    ).scalar() or 0
    total_expenses = query.filter(Transaction.amount < 0).with_entities(
        func.sum(Transaction.amount)
    ).scalar() or 0

    # Get category breakdown (use subcategory column)
    category_breakdown = query.with_entities(
        Transaction.subcategory,
        func.count(Transaction.id).label("count"),
        func.sum(Transaction.amount).label("total")
    ).group_by(Transaction.subcategory).all()
    
    categories = [
        CategorySummary(
            category=cat[0] or "Uncategorized",
            transaction_count=cat[1],
            total_amount=float(cat[2]) if cat[2] else 0
        )
        for cat in category_breakdown
    ]

    # Get monthly trend
    monthly_trend = query.with_entities(
        extract('year', Transaction.date).label('year'),
        extract('month', Transaction.date).label('month'),
        func.sum(Transaction.amount).label('total')
    ).group_by('year', 'month').order_by('year', 'month').all()

    return TransactionStats(
        total_transactions=total_count,
        total_income=float(total_income) if total_income else 0,
        total_expenses=abs(float(total_expenses)) if total_expenses else 0,
        net_amount=float(total_income + total_expenses) if (total_income and total_expenses) else 0,
        categories=categories,
        monthly_trend=[
            {
                "month": f"{int(m[0])}-{int(m[1]):02d}",
                "amount": float(m[2]) if m[2] else 0
            }
            for m in monthly_trend
        ]
    )


@router.get("/{transaction_id}", response_model=TransactionSchema)
async def get_transaction(
    transaction_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_verified_user)
):
    """
    Get a specific transaction by ID.
    """
    transaction = db.query(Transaction).join(Account).filter(
        Transaction.id == transaction_id,
        Account.user_id == current_user.id
    ).first()
    
    if not transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    return transaction


@router.put("/{transaction_id}", response_model=TransactionSchema)
async def update_transaction(
    transaction_id: UUID,
    update_data: TransactionUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_verified_user)
):
    """
    Update a transaction's details.
    """
    transaction = db.query(Transaction).join(Account).filter(
        Transaction.id == transaction_id,
        Account.user_id == current_user.id
    ).first()
    
    if not transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    # Update fields
    update_dict = update_data.dict(exclude_unset=True)
    for field, value in update_dict.items():
        if hasattr(transaction, field):
            # Handle amount conversion
            if field == "amount":
                setattr(transaction, "amount_cents", int(value * 100))
            else:
                setattr(transaction, field, value)
    
    # Update timestamp
    transaction.updated_at = datetime.utcnow()
    
    db.commit()
    db.refresh(transaction)
    
    return transaction


@router.post("/bulk", response_model=Dict[str, Any])
async def bulk_update_transactions(
    bulk_update: BulkTransactionUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_verified_user)
):
    """
    Perform bulk operations on multiple transactions.
    """
    # Verify all transactions belong to user
    transactions = db.query(Transaction).join(Account).filter(
        Transaction.id.in_(bulk_update.transaction_ids),
        Account.user_id == current_user.id
    ).all()
    
    if len(transactions) != len(bulk_update.transaction_ids):
        raise HTTPException(status_code=403, detail="Some transactions not found or unauthorized")
    
    updated_count = 0
    
    # Apply updates based on operation
    if bulk_update.operation == "categorize":
        for transaction in transactions:
            transaction.user_category = bulk_update.category
            transaction.updated_at = datetime.utcnow()
            updated_count += 1
    
    elif bulk_update.operation == "reconcile":
        for transaction in transactions:
            transaction.is_reconciled = True
            transaction.updated_at = datetime.utcnow()
            updated_count += 1
    
    elif bulk_update.operation == "unreconcile":
        for transaction in transactions:
            transaction.is_reconciled = False
            transaction.updated_at = datetime.utcnow()
            updated_count += 1
    
    elif bulk_update.operation == "add_tag":
        for transaction in transactions:
            current_tags = transaction.tags or []
            if bulk_update.tag and bulk_update.tag not in current_tags:
                current_tags.append(bulk_update.tag)
                transaction.tags = current_tags
                transaction.updated_at = datetime.utcnow()
                updated_count += 1
    
    elif bulk_update.operation == "remove_tag":
        for transaction in transactions:
            current_tags = transaction.tags or []
            if bulk_update.tag and bulk_update.tag in current_tags:
                current_tags.remove(bulk_update.tag)
                transaction.tags = current_tags
                transaction.updated_at = datetime.utcnow()
                updated_count += 1
    
    elif bulk_update.operation == "delete":
        for transaction in transactions:
            db.delete(transaction)
            updated_count += 1
    
    db.commit()
    
    return {
        "success": True,
        "updated_count": updated_count,
        "operation": bulk_update.operation
    }


@router.get("/export/csv")
async def export_transactions_csv(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    account_id: Optional[UUID] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_verified_user)
):
    """
    Export transactions as CSV file.
    """
    # Build query
    query = db.query(Transaction).join(Account).filter(Account.user_id == current_user.id)
    
    if start_date:
        query = query.filter(Transaction.date >= start_date)
    if end_date:
        query = query.filter(Transaction.date <= end_date)
    if account_id:
        query = query.filter(Transaction.account_id == account_id)
    
    transactions = query.order_by(Transaction.date.desc()).all()
    
    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow([
        "Date", "Description", "Merchant", "Amount", "Category",
        "Account", "Status", "Reconciled", "Notes", "Tags"
    ])
    
    # Write transactions
    for txn in transactions:
        writer.writerow([
            txn.date.strftime("%Y-%m-%d"),
            txn.name,
            txn.merchant_name or "",
            f"{txn.amount:.2f}",
            txn.user_category or txn.primary_category or "",
            txn.account.name,
            "Pending" if txn.pending else "Posted",
            "Yes" if txn.is_reconciled else "No",
            txn.notes or "",
            ", ".join(txn.tags) if txn.tags else ""
        ])
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode()),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=transactions_{datetime.now().strftime('%Y%m%d')}.csv"
        }
    )


@router.get("/export/excel")
async def export_transactions_excel(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    account_id: Optional[UUID] = Query(None),
    include_summary: bool = Query(True, description="Include summary sheet"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_verified_user)
):
    """
    Export transactions as Excel file with formatting and optional summary.
    """
    # Build query
    query = db.query(Transaction).join(Account).filter(Account.user_id == current_user.id)
    
    if start_date:
        query = query.filter(Transaction.date >= start_date)
    if end_date:
        query = query.filter(Transaction.date <= end_date)
    if account_id:
        query = query.filter(Transaction.account_id == account_id)
    
    transactions = query.order_by(Transaction.date.desc()).all()
    
    # Create workbook
    wb = Workbook()
    
    # Transactions sheet
    ws = wb.active
    ws.title = "Transactions"
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        "Date", "Description", "Merchant", "Amount", "Category",
        "Account", "Status", "Reconciled", "Notes", "Tags"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Data rows
    for row, txn in enumerate(transactions, 2):
        ws.cell(row=row, column=1, value=txn.date)
        ws.cell(row=row, column=2, value=txn.name)
        ws.cell(row=row, column=3, value=txn.merchant_name or "")
        ws.cell(row=row, column=4, value=txn.amount)
        ws.cell(row=row, column=5, value=txn.user_category or txn.primary_category or "")
        ws.cell(row=row, column=6, value=txn.account.name)
        ws.cell(row=row, column=7, value="Pending" if txn.pending else "Posted")
        ws.cell(row=row, column=8, value="Yes" if txn.is_reconciled else "No")
        ws.cell(row=row, column=9, value=txn.notes or "")
        ws.cell(row=row, column=10, value=", ".join(txn.tags) if txn.tags else "")
        
        # Format amount cell
        amount_cell = ws.cell(row=row, column=4)
        amount_cell.number_format = "#,##0.00"
        if txn.amount < 0:
            amount_cell.font = Font(color="FF0000")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add summary sheet if requested
    if include_summary:
        summary_ws = wb.create_sheet(title="Summary")
        
        # Calculate summary statistics
        total_income = sum(t.amount for t in transactions if t.amount > 0)
        total_expenses = sum(t.amount for t in transactions if t.amount < 0)
        net_amount = total_income + total_expenses
        
        # Category breakdown
        category_totals = {}
        for txn in transactions:
            cat = txn.user_category or txn.primary_category or "Uncategorized"
            if cat not in category_totals:
                category_totals[cat] = 0
            category_totals[cat] += txn.amount
        
        # Write summary
        summary_data = [
            ["Transaction Summary", ""],
            ["", ""],
            ["Period", f"{start_date or 'All time'} to {end_date or 'Present'}"],
            ["Total Transactions", len(transactions)],
            ["", ""],
            ["Financial Summary", ""],
            ["Total Income", total_income],
            ["Total Expenses", abs(total_expenses)],
            ["Net Amount", net_amount],
            ["", ""],
            ["Category Breakdown", "Amount"],
        ]
        
        for cat, amount in sorted(category_totals.items(), key=lambda x: abs(x[1]), reverse=True):
            summary_data.append([cat, amount])
        
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = summary_ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Format headers
                if row_idx in [1, 6, 11]:
                    cell.font = Font(bold=True, size=12)
                
                # Format numbers
                if col_idx == 2 and isinstance(value, (int, float)):
                    cell.number_format = "#,##0.00"
                    if value < 0:
                        cell.font = Font(color="FF0000")
        
        # Auto-adjust summary columns
        for column in summary_ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            summary_ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=transactions_{datetime.now().strftime('%Y%m%d')}.xlsx"
        }
    )


@router.post("/import/csv")
async def import_transactions_csv(
    file: UploadFile = File(...),
    account_id: UUID = Query(..., description="Account to import transactions to"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_verified_user)
):
    """
    Import transactions from CSV file.
    """
    # Verify account belongs to user
    account = db.query(Account).filter(
        Account.id == account_id,
        Account.user_id == current_user.id
    ).first()
    
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    # Read CSV
    contents = await file.read()
    csv_reader = csv.DictReader(io.StringIO(contents.decode()))
    
    imported_count = 0
    errors = []
    
    for row_num, row in enumerate(csv_reader, 2):
        try:
            # Parse transaction data
            transaction = Transaction(
                account_id=account_id,
                plaid_transaction_id=f"manual_{account_id}_{datetime.now().timestamp()}_{row_num}",
                date=datetime.strptime(row.get("Date", ""), "%Y-%m-%d").date(),
                name=row.get("Description", ""),
                merchant_name=row.get("Merchant", ""),
                amount_cents=int(float(row.get("Amount", 0)) * 100),
                user_category=row.get("Category", ""),
                notes=row.get("Notes", ""),
                is_reconciled=row.get("Reconciled", "").lower() in ["yes", "true", "1"],
                tags=row.get("Tags", "").split(",") if row.get("Tags") else []
            )
            
            db.add(transaction)
            imported_count += 1
            
        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")
    
    db.commit()
    
    return {
        "success": True,
        "imported_count": imported_count,
        "errors": errors
    }